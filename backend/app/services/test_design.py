import uuid
import asyncio
import logging
from typing import List, Optional, Dict, Any
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, update, func, and_, or_
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.agents.prompts import PromptTemplates
from app.models.db_models import Requirement, SplitRequirement, TestPoint, TestCase, AISession, AIMessage, Task

logger = logging.getLogger("uvicorn.error")
from app.models.test_design import (
    RequirementListItem, RequirementListResponse,
    ImportRequirementRequest, ImportRequirementResponse,
    MindMapNode, MindMapNodeData,
    TestPointCreate, TestPointUpdate, TestPointResponse,
    TestCaseCreate, TestCaseUpdate, TestCaseResponse, TestCaseStep,
    AIAdjustStart, AIAdjustApply, AIAdjustApplyResponse,
    GenerateResponse, TaskStatusResponse,
    ResponseModel,
    AdoptProposalResponse, RejectProposalResponse,
)


class TestDesignService:
    def __init__(self):
        from app.agents.llm_client import LLMClient
        self.tasks: Dict[str, asyncio.Task] = {}
        self._llm_client = LLMClient()
        self._llm_available = bool(settings.OPENAI_API_KEY)

    # ========== 需求列表 ==========
    async def import_requirement(
        self, db: AsyncSession, data: ImportRequirementRequest
    ) -> ImportRequirementResponse:
        req_id = f"req-{uuid.uuid4().hex[:8]}"
        now = datetime.utcnow()

        requirement = Requirement(
            id=req_id,
            user_id="00000000-0000-0000-0000-000000000000",
            title=data.title,
            content=data.standardizedContent,
            status="confirmed",
            source="standardization",
            created_at=now,
            updated_at=now,
        )
        db.add(requirement)

        for idx, sr in enumerate(data.splitRequirements):
            if not sr.selected:
                continue
            sr_id = f"sr-{uuid.uuid4().hex[:8]}"
            split_req = SplitRequirement(
                id=sr_id,
                requirement_id=req_id,
                text=sr.content,
                status="pending",
                sort_order=idx,
                created_at=now,
            )
            db.add(split_req)

        await db.commit()

        return ImportRequirementResponse(
            id=req_id,
            title=data.title,
            status="confirmed",
            statusText="待生成测试点",
            date=now.strftime("%Y-%m-%d %H:%M"),
            testPointCount=0,
            caseCount=0,
            source="standardization",
        )

    async def get_requirements_list(
        self, db: AsyncSession, page: int, pageSize: int, status: Optional[str], keyword: Optional[str]
    ) -> RequirementListResponse:
        # 基础过滤：直接使用 Requirement.status，不再关联 Task 表
        base_filters = [
            Requirement.status.in_([
                "confirmed", "generating_points", "points_generated",
                "generating_cases", "completed", "failed", "pending"
            ]),
        ]
        if keyword:
            base_filters.append(Requirement.title.contains(keyword))

        base_query = select(Requirement).where(and_(*base_filters)).order_by(Requirement.updated_at.desc())
        result = await db.execute(base_query)
        all_requirements = result.scalars().all()

        # 对于 generating_points / generating_cases 状态，批量查询对应 running task 的进度
        req_ids = [r.id for r in all_requirements]
        req_task_progress = {}
        if req_ids:
            task_subq = (
                select(
                    Task.requirement_id,
                    Task.progress,
                    func.row_number().over(
                        partition_by=Task.requirement_id,
                        order_by=Task.created_at.desc()
                    ).label('rn')
                ).where(
                    and_(
                        Task.requirement_id.in_(req_ids),
                        Task.status.in_(["running", "pending"])
                    )
                )
            ).subquery()
            task_query = select(task_subq.c.requirement_id, task_subq.c.progress).where(task_subq.c.rn == 1)
            task_result = await db.execute(task_query)
            for row in task_result:
                req_task_progress[row[0]] = row[1]

        # status -> statusText 映射
        status_text_map = {
            "confirmed": "待生成测试点",
            "pending": "待生成测试点",
            "generating_points": "测试点生成中",
            "points_generated": "测试点已生成",
            "generating_cases": "用例生成中",
            "completed": "已完成",
            "failed": "生成失败",
        }

        # 状态筛选映射
        status_mapping = {
            "pending": "confirmed",
            "confirmed": "confirmed",
            "generating": ["generating_points", "generating_cases"],
            "points_generated": "points_generated",
            "completed": "completed",
            "failed": "failed",
        }

        all_items = []
        for req in all_requirements:
            st = req.status or "confirmed"
            progress = req_task_progress.get(req.id, 0)
            display_text = status_text_map.get(st, "待生成测试点")
            if st in ("generating_points", "generating_cases") and progress > 0:
                display_text += f" {progress}%"

            # 状态筛选
            if status:
                mapped = status_mapping.get(status)
                if isinstance(mapped, list):
                    if st not in mapped:
                        continue
                elif mapped and st != mapped:
                    continue
                elif not mapped and st != status:
                    continue

            tp_count = await self._get_test_point_count(db, req.id)
            case_count = await self._get_case_count(db, req.id)
            all_items.append(RequirementListItem(
                id=req.id,
                title=req.title,
                status=st,
                statusText=display_text,
                date=req.updated_at.isoformat() + "Z" if req.updated_at else "",
                testPointCount=tp_count,
                caseCount=case_count,
                source=req.source or "standardization"
            ))

        # 分页
        total = len(all_items)
        start = (page - 1) * pageSize
        end = start + pageSize
        paged_items = all_items[start:end]

        return RequirementListResponse(list=paged_items, total=total, page=page, pageSize=pageSize)

    async def _get_test_point_count(self, db: AsyncSession, requirement_id: str) -> int:
        query = select(func.count(TestPoint.id)).join(SplitRequirement).where(
            SplitRequirement.requirement_id == requirement_id
        )
        result = await db.execute(query)
        return result.scalar() or 0

    async def _get_case_count(self, db: AsyncSession, requirement_id: str) -> int:
        query = select(func.count(TestCase.id)).join(TestPoint).join(SplitRequirement).where(
            SplitRequirement.requirement_id == requirement_id
        )
        result = await db.execute(query)
        return result.scalar() or 0

    # ========== 脑图数据 ==========
    async def get_mindmap_data(self, db: AsyncSession, requirement_id: str) -> MindMapNode:
        result = await db.execute(
            select(Requirement).where(Requirement.id == requirement_id)
        )
        requirement = result.scalar_one_or_none()
        if not requirement:
            return MindMapNode(
                data=MindMapNodeData(id="", text="", level="root", status="pending"),
                children=[]
            )
        
        sr_result = await db.execute(
            select(SplitRequirement)
            .where(SplitRequirement.requirement_id == requirement_id)
            .order_by(SplitRequirement.sort_order)
            .options(selectinload(SplitRequirement.test_points).selectinload(TestPoint.test_cases))
        )
        split_reqs = sr_result.scalars().all()
        
        children = []
        for sr in split_reqs:
            tp_children = []
            for tp in sr.test_points:
                case_children = []
                for tc in tp.test_cases:
                    note_html = self._build_case_note_html(tc)
                    case_children.append(MindMapNode(
                        data=MindMapNodeData(
                            id=tc.id,
                            text=tc.text,
                            level="testCase",
                            case_property=tc.case_property,
                            source=tc.source,
                            note=note_html,
                            pre_condition=tc.pre_condition,
                            steps=tc.steps,
                            marked=tc.marked
                        ),
                        children=[]
                    ))
                tp_children.append(MindMapNode(
                    data=MindMapNodeData(
                        id=tp.id,
                        text=tp.text,
                        level="testPoint",
                        status=tp.status,
                        source=tp.source,
                        marked=tp.marked,
                        description=tp.description
                    ),
                    children=case_children
                ))
            children.append(MindMapNode(
                data=MindMapNodeData(
                    id=sr.id,
                    text=sr.text,
                    level="requirement",
                    status=sr.status
                ),
                children=tp_children
            ))
        
        return MindMapNode(
            data=MindMapNodeData(
                id=requirement.id,
                text=requirement.title,
                level="root",
                status=requirement.status
            ),
            children=children
        )

    def _build_case_note_html(self, tc: TestCase) -> str:
        steps_html = ""
        if tc.steps:
            for step in tc.steps:
                steps_html += f"<div class='step'><b>{step.get('name', '')}</b>: {step.get('description', '')} → {step.get('stepExpectedResult', '')}</div>"
        return f"<div class='case-note-popover'><p><b>前置条件:</b> {tc.pre_condition or '无'}</p><p><b>步骤:</b></p>{steps_html}</div>"

    def _build_case_note_html_from_data(self, pre_condition: str, steps: list) -> str:
        steps_html = ""
        if steps:
            for step in steps:
                steps_html += f"<div class='step'><b>{step.get('name', '')}</b>: {step.get('description', '')} → {step.get('stepExpectedResult', '')}</div>"
        return f"<div class='case-note-popover'><p><b>前置条件:</b> {pre_condition or '无'}</p><p><b>步骤:</b></p>{steps_html}</div>"

    # ========== 测试点管理 ==========
    async def create_test_point(self, db: AsyncSession, requirement_id: str, data: TestPointCreate) -> TestPointResponse:
        test_point = TestPoint(
            split_requirement_id=data.requirementNodeId,
            text=data.text,
            description=data.description,
            source="人工"
        )
        db.add(test_point)
        await db.commit()
        await db.refresh(test_point)
        return TestPointResponse(id=test_point.id, text=test_point.text, _source="人工")

    async def update_test_point(self, db: AsyncSession, test_point_id: str, data: TestPointUpdate) -> TestPointResponse:
        result = await db.execute(
            update(TestPoint).where(TestPoint.id == test_point_id).values(text=data.text)
        )
        await db.commit()
        return TestPointResponse(id=test_point_id, text=data.text, _source="人工")

    async def delete_test_point(self, db: AsyncSession, test_point_id: str) -> bool:
        await db.execute(delete(TestPoint).where(TestPoint.id == test_point_id))
        await db.commit()
        return True

    async def batch_delete_test_points(self, db: AsyncSession, ids: List[str]) -> bool:
        await db.execute(delete(TestPoint).where(TestPoint.id.in_(ids)))
        await db.commit()
        return True

    async def mark_test_point(self, db: AsyncSession, test_point_id: str, marked: bool) -> bool:
        await db.execute(
            update(TestPoint).where(TestPoint.id == test_point_id).values(marked=marked)
        )
        await db.commit()
        return True

    # ========== 测试用例管理 ==========
    async def create_test_case(self, db: AsyncSession, test_point_id: str, data: TestCaseCreate) -> TestCaseResponse:
        steps = []
        if data.steps:
            steps = [step.model_dump() for step in data.steps]
        test_case = TestCase(
            test_point_id=test_point_id,
            text=data.text,
            case_property=data.caseProperty,
            pre_condition=data.preCondition,
            steps=steps,
            source="人工"
        )
        db.add(test_case)
        await db.commit()
        await db.refresh(test_case)
        return TestCaseResponse(
            id=test_case.id,
            text=test_case.text,
            caseProperty=test_case.case_property,
            preCondition=test_case.pre_condition,
            steps=data.steps
        )

    async def update_test_case(self, db: AsyncSession, test_case_id: str, data: TestCaseUpdate) -> TestCaseResponse:
        steps = []
        if data.steps:
            steps = [step.model_dump() for step in data.steps]
        await db.execute(
            update(TestCase).where(TestCase.id == test_case_id).values(
                text=data.text,
                case_property=data.caseProperty,
                pre_condition=data.preCondition,
                steps=steps
            )
        )
        await db.commit()
        return TestCaseResponse(
            id=test_case_id,
            text=data.text,
            caseProperty=data.caseProperty,
            preCondition=data.preCondition,
            steps=data.steps
        )

    async def delete_test_case(self, db: AsyncSession, test_case_id: str) -> bool:
        await db.execute(delete(TestCase).where(TestCase.id == test_case_id))
        await db.commit()
        return True

    async def batch_delete_test_cases(self, db: AsyncSession, ids: List[str]) -> bool:
        await db.execute(delete(TestCase).where(TestCase.id.in_(ids)))
        await db.commit()
        return True

    async def mark_test_case(self, db: AsyncSession, test_case_id: str, marked: bool) -> bool:
        await db.execute(
            update(TestCase).where(TestCase.id == test_case_id).values(marked=marked)
        )
        await db.commit()
        return True

    # ========== AI调整 ==========
    async def start_ai_session(self, db: AsyncSession, data: AIAdjustStart) -> Dict[str, Any]:
        session = AISession(
            requirement_id=data.requirementId,
            node_id=data.nodeId,
            node_type=data.nodeType,
            marked_node_ids=data.markedNodeIds or []
        )
        db.add(session)
        await db.commit()
        await db.refresh(session)

        context_data = await self._get_ai_adjust_context(db, data)
        system_prompt = self._build_ai_adjust_prompt(data.nodeType, data.markedNodeIds or [], context_data)
        ai_message = AIMessage(session_id=session.id, role="system", content=system_prompt)
        db.add(ai_message)
        await db.commit()

        return {"sessionId": session.id, "message": "AI调整会话已创建"}

    async def _get_ai_adjust_context(self, db: AsyncSession, data: AIAdjustStart) -> Dict[str, Any]:
        if data.nodeType == "requirement":
            sr_result = await db.execute(
                select(SplitRequirement).where(SplitRequirement.id == data.nodeId)
            )
            sr = sr_result.scalar_one_or_none()
            tp_result = await db.execute(
                select(TestPoint)
                .where(TestPoint.split_requirement_id == data.nodeId)
                .order_by(TestPoint.created_at)
            )
            test_points = tp_result.scalars().all()
            return {
                "node_text": sr.text if sr else "",
                "existing_items": [{"id": tp.id, "text": tp.text, "marked": tp.marked} for tp in test_points],
                "item_label": "测试点",
            }
        else:
            tp_result = await db.execute(
                select(TestPoint).where(TestPoint.id == data.nodeId)
            )
            tp = tp_result.scalar_one_or_none()
            tc_result = await db.execute(
                select(TestCase)
                .where(TestCase.test_point_id == data.nodeId)
                .order_by(TestCase.created_at)
            )
            test_cases = tc_result.scalars().all()
            return {
                "node_text": tp.text if tp else "",
                "existing_items": [
                    {"id": tc.id, "text": tc.text, "property": tc.case_property, "marked": tc.marked}
                    for tc in test_cases
                ],
                "item_label": "测试用例",
            }

    def _build_ai_adjust_prompt(self, node_type: str, marked_node_ids: List[str], context: Dict[str, Any]) -> str:
        node_text = context.get("node_text", "")
        existing_items = context.get("existing_items", [])
        item_label = context.get("item_label", "")

        items_text = ""
        if existing_items:
            for item in existing_items:
                marker = " [标记保留]" if item.get("marked") else ""
                prop = f" [{item.get('property', '')}]" if item.get("property") else ""
                item_id = item.get("id", "")
                items_text += f"  - [ID: {item_id}] {item['text']}{prop}{marker}\n"
        else:
            items_text = "  （暂无）\n"

        if node_type == "requirement":
            return (
                "你是一个专业的测试设计专家。以下是当前需求的拆分内容和已有的测试点，"
                "请基于这些信息帮助用户调整、补充或重新生成测试点。\n"
                f"\n【当前需求拆分内容】\n{node_text}\n"
                f"\n【已有{item_label}】\n{items_text}\n"
                "【标记保留的测试点ID】\n"
                f"{', '.join(marked_node_ids) if marked_node_ids else '无'}\n"
                "\n注意：标记保留的测试点不可删除或修改其内容。"
                "请根据用户的调整要求，在保留已有有效测试点的基础上，补充或优化测试点。"
                "\n\n当你给出调整建议(type=proposal)时，必须在pending_nodes中列出所有变更："
                "\n- 新增测试点：action为add，填写text和可选的description"
                "\n- 删除测试点：action为remove，填写id为已有测试点的ID（不可删除标记保留的测试点）"
                "\n- 修改测试点：action为modify，填写id为已有测试点的ID，并填写需要修改的字段（text、description等）"
                "\n- 重要：已有节点列表中每个节点前标注了[ID: xxx]，修改或删除时必须使用该ID"
            )
        else:
            return (
                "你是一个专业的测试设计专家。以下是当前测试点的内容和已有的测试用例，"
                "请基于这些信息帮助用户调整、补充或重新生成测试用例（包含正例和反例）。\n"
                f"\n【当前测试点内容】\n{node_text}\n"
                f"\n【已有{item_label}】\n{items_text}\n"
                "【标记保留的测试用例ID】\n"
                f"{', '.join(marked_node_ids) if marked_node_ids else '无'}\n"
                "\n注意：标记保留的测试用例不可删除或修改其内容。"
                "请根据用户的调整要求，在保留已有有效测试用例的基础上，补充或优化测试用例。"
                "每个测试用例需包含：用例名称、用例属性（正例/反例）、前置条件、测试步骤。"
                "\n\n当你给出调整建议(type=proposal)时，必须在pending_nodes中列出所有变更："
                "\n- 新增测试用例：action为add，填写text、case_property（正例/反例）、可选的pre_condition和steps"
                "\n- 删除测试用例：action为remove，填写id为已有测试用例的ID（不可删除标记保留的测试用例）"
                "\n- 修改测试用例：action为modify，填写id为已有测试用例的ID，并填写需要修改的字段（text、case_property、pre_condition、steps等）"
                "\n- 重要：已有节点列表中每个节点前标注了[ID: xxx]，修改或删除时必须使用该ID"
            )

    async def _call_llm_with_schema(self, messages, schema_description, temperature=0.7, max_tokens=8192):
        if not self._llm_available:
            logger.warning("LLM call with schema skipped: API key not configured")
            return None
        try:
            return await self._llm_client.chat_with_schema(
                messages=messages,
                schema_description=schema_description,
                temperature=temperature,
                max_tokens=max_tokens
            )
        except Exception as e:
            logger.warning(f"LLM call with schema failed: {type(e).__name__}: {e}")
            return None

    async def send_ai_message(self, db: AsyncSession, session_id: str, content: str, marked_node_ids: Optional[List[str]] = None) -> Dict[str, Any]:
        # 如果有传入最新的标记数据，更新会话的 marked_node_ids
        if marked_node_ids is not None:
            session_update_result = await db.execute(
                select(AISession).where(AISession.id == session_id)
            )
            session_obj = session_update_result.scalar_one_or_none()
            if session_obj:
                session_obj.marked_node_ids = marked_node_ids
                await db.commit()

        user_msg = AIMessage(session_id=session_id, role="user", content=content, msg_type="text")
        db.add(user_msg)
        await db.commit()

        messages_result = await db.execute(
            select(AIMessage).where(AIMessage.session_id == session_id).order_by(AIMessage.created_at)
        )
        messages = messages_result.scalars().all()

        # 获取会话信息，为本次 LLM 调用构建最新的上下文提示词（内存中构建，不修改 DB 消息记录）
        session_info_result = await db.execute(
            select(AISession).where(AISession.id == session_id)
        )
        session_info = session_info_result.scalar_one_or_none()
        latest_marked_ids = (marked_node_ids if marked_node_ids is not None
                             else (session_info.marked_node_ids if session_info else []))

        context_data = await self._get_ai_adjust_context(db, AIAdjustStart(
            requirementId=session_info.requirement_id if session_info else "",
            nodeId=session_info.node_id if session_info else "",
            nodeType=session_info.node_type if session_info else "requirement",
            markedNodeIds=latest_marked_ids
        ))
        fresh_system_prompt = self._build_ai_adjust_prompt(
            session_info.node_type if session_info else "requirement",
            latest_marked_ids,
            context_data
        )

        # 构建 API 消息列表：system 消息使用最新构建的提示词，其他消息取 DB 中的历史记录
        api_messages = []
        system_replaced = False
        for msg in messages:
            if msg.role == "system":
                api_messages.append({"role": "system", "content": fresh_system_prompt})
                system_replaced = True
            else:
                api_messages.append({"role": msg.role, "content": msg.content})
        if not system_replaced:
            api_messages.insert(0, {"role": "system", "content": fresh_system_prompt})

        try:
            ai_result = await self._call_llm_with_schema(
                messages=api_messages,
                schema_description=PromptTemplates.TEST_DESIGN_ADJUST_SCHEMA,
                temperature=0.7,
                max_tokens=8192
            )

            if not ai_result:
                ai_result = {
                    "content": "AI服务暂时不可用，请稍后重试。",
                    "type": "discussion",
                    "change_summary": ""
                }

            ai_content = ai_result.get("content", "")
            msg_type = ai_result.get("type", "discussion")
            change_summary = ai_result.get("change_summary") if msg_type == "proposal" else None
            pending_nodes = ai_result.get("pending_nodes") if msg_type == "proposal" else None

            pending_mindmap_data = None
            if msg_type == "proposal" and pending_nodes:
                session_result = await db.execute(
                    select(AISession).where(AISession.id == session_id)
                )
                session_obj2 = session_result.scalar_one_or_none()
                if session_obj2:
                    node_data = await self._build_pending_mindmap_data(
                        db, session_obj2.node_id, session_obj2.node_type,
                        session_obj2.marked_node_ids or [], pending_nodes
                    )
                    if node_data:
                        pending_mindmap_data = {
                            "nodeData": node_data,
                            "adjustNodes": pending_nodes
                        }

            assistant_msg = AIMessage(
                session_id=session_id, role="assistant", content=ai_content,
                msg_type=msg_type, change_summary=change_summary,
                pending_mindmap_data=pending_mindmap_data
            )
        except Exception as e:
            ai_content = f"AI服务暂时不可用，请稍后重试。错误: {str(e)}"
            assistant_msg = AIMessage(
                session_id=session_id, role="assistant", content=ai_content,
                msg_type="text"
            )

        db.add(assistant_msg)
        await db.commit()
        await db.refresh(assistant_msg)

        return {
            "id": assistant_msg.id,
            "role": "assistant",
            "content": ai_content,
            "type": assistant_msg.msg_type or "text",
            "changeSummary": assistant_msg.change_summary,
            "pendingMindMapData": pending_mindmap_data.get("nodeData") if isinstance(pending_mindmap_data, dict) and "nodeData" in pending_mindmap_data else None,
            "timestamp": assistant_msg.created_at.isoformat() + "Z" if assistant_msg.created_at else None,
        }

    async def adopt_proposal(self, db: AsyncSession, session_id: str, message_id: str, requirement_id: str) -> AdoptProposalResponse:
        result = await db.execute(
            select(AIMessage).where(
                and_(AIMessage.id == message_id, AIMessage.session_id == session_id)
            )
        )
        msg = result.scalar_one_or_none()
        if not msg:
            raise ValueError("消息不存在")

        msg.adopted = True
        msg.rejected = False

        session_result = await db.execute(
            select(AISession).where(AISession.id == session_id)
        )
        session = session_result.scalar_one_or_none()
        await db.commit()

        if msg.pending_mindmap_data and session:
            new_node_mappings = await self._apply_mindmap_changes(
                db, session.requirement_id, session.node_id, session.node_type,
                session.marked_node_ids or [], msg.pending_mindmap_data
            )
        else:
            new_node_mappings = []

        return AdoptProposalResponse(messageId=message_id, adopted=True, newNodeMappings=new_node_mappings)

    async def reject_proposal(self, db: AsyncSession, session_id: str, message_id: str, requirement_id: str) -> RejectProposalResponse:
        result = await db.execute(
            select(AIMessage).where(
                and_(AIMessage.id == message_id, AIMessage.session_id == session_id)
            )
        )
        msg = result.scalar_one_or_none()
        if not msg:
            raise ValueError("消息不存在")

        msg.rejected = True
        msg.adopted = False
        await db.commit()

        return RejectProposalResponse(messageId=message_id, rejected=True)

    def _extract_change_summary(self, content: str) -> Optional[str]:
        for line in content.split("\n"):
            line = line.strip()
            if line.startswith("变更摘要") or line.startswith("调整摘要") or line.startswith("变更内容"):
                return line.split("：", 1)[-1].split(":", 1)[-1].strip() if ("：" in line or ":" in line) else line
        lines = content.strip().split("\n")
        first_line = lines[0] if lines else ""
        if len(first_line) <= 120:
            return first_line
        return content[:120] + "..."

    async def _build_pending_mindmap_data(
        self, db: AsyncSession, node_id: str, node_type: str,
        marked_node_ids: List[str], pending_nodes: List[Dict]
    ) -> Optional[Dict[str, Any]]:
        if node_type == "requirement":
            sr_result = await db.execute(
                select(SplitRequirement).where(SplitRequirement.id == node_id)
            )
            sr = sr_result.scalar_one_or_none()
            if not sr:
                return None

            tp_result = await db.execute(
                select(TestPoint)
                .where(TestPoint.split_requirement_id == node_id)
                .order_by(TestPoint.created_at)
            )
            test_points = tp_result.scalars().all()

            children = []
            for tp in test_points:
                children.append({
                    "data": {
                        "id": tp.id,
                        "text": tp.text,
                        "_level": "testPoint",
                        "_status": tp.status,
                        "_source": tp.source,
                        "_marked": tp.marked,
                        "description": tp.description
                    },
                    "children": []
                })

            remove_ids = set()
            modify_map = {}
            for node in pending_nodes:
                action = node.get("action", "")
                if action == "add":
                    children.append({
                        "data": {
                            "text": node.get("text", ""),
                            "_level": "testPoint",
                            "_source": "AI",
                            "_marked": False,
                            "description": node.get("description")
                        },
                        "children": []
                    })
                elif action == "remove":
                    target_id = node.get("id", "")
                    if target_id and target_id not in marked_node_ids:
                        remove_ids.add(target_id)
                elif action == "modify":
                    target_id = node.get("id", "")
                    if target_id and target_id not in marked_node_ids:
                        modify_map[target_id] = node

            # 根据 marked_node_ids 覆盖 _marked 状态（对话过程中标记的数据未回写DB）
            for child in children:
                child_id = child["data"].get("id")
                if child_id and child_id in marked_node_ids:
                    child["data"]["_marked"] = True

            if remove_ids:
                children = [c for c in children if c["data"].get("id") not in remove_ids]

            for child in children:
                child_id = child["data"].get("id")
                if child_id and child_id in modify_map:
                    mod = modify_map[child_id]
                    if mod.get("text"):
                        child["data"]["text"] = mod["text"]
                    if mod.get("description") is not None:
                        child["data"]["description"] = mod["description"]

            return {
                "data": {
                    "id": sr.id,
                    "text": sr.text,
                    "_level": "requirement",
                    "_status": sr.status
                },
                "children": children
            }

        else:
            tp_result = await db.execute(
                select(TestPoint).where(TestPoint.id == node_id)
            )
            tp = tp_result.scalar_one_or_none()
            if not tp:
                return None

            tc_result = await db.execute(
                select(TestCase)
                .where(TestCase.test_point_id == node_id)
                .order_by(TestCase.created_at)
            )
            test_cases = tc_result.scalars().all()

            children = []
            for tc in test_cases:
                note_html = self._build_case_note_html(tc)
                children.append({
                    "data": {
                        "id": tc.id,
                        "text": tc.text,
                        "_level": "testCase",
                        "_caseProperty": tc.case_property,
                        "_source": tc.source,
                        "_marked": tc.marked,
                        "note": note_html,
                        "_preCondition": tc.pre_condition,
                        "steps": tc.steps
                    },
                    "children": []
                })

            remove_ids = set()
            modify_map = {}
            for node in pending_nodes:
                action = node.get("action", "")
                if action == "add":
                    children.append({
                        "data": {
                            "text": node.get("text", ""),
                            "_level": "testCase",
                            "_caseProperty": node.get("case_property", "正例"),
                            "_source": "AI",
                            "_marked": False,
                            "_preCondition": node.get("pre_condition"),
                            "steps": node.get("steps", [])
                        },
                        "children": []
                    })
                elif action == "remove":
                    target_id = node.get("id", "")
                    if target_id and target_id not in marked_node_ids:
                        remove_ids.add(target_id)
                elif action == "modify":
                    target_id = node.get("id", "")
                    if target_id and target_id not in marked_node_ids:
                        modify_map[target_id] = node

            # 根据 marked_node_ids 覆盖 _marked 状态（对话过程中标记的数据未回写DB）
            for child in children:
                child_id = child["data"].get("id")
                if child_id and child_id in marked_node_ids:
                    child["data"]["_marked"] = True

            if remove_ids:
                children = [c for c in children if c["data"].get("id") not in remove_ids]

            for child in children:
                child_id = child["data"].get("id")
                if child_id and child_id in modify_map:
                    mod = modify_map[child_id]
                    if mod.get("text"):
                        child["data"]["text"] = mod["text"]
                    if mod.get("case_property"):
                        child["data"]["_caseProperty"] = mod["case_property"]
                    if mod.get("pre_condition") is not None:
                        child["data"]["_preCondition"] = mod["pre_condition"]
                    if mod.get("steps") is not None:
                        child["data"]["steps"] = mod["steps"]
                        note_html = self._build_case_note_html_from_data(mod.get("pre_condition", ""), mod.get("steps", []))
                        child["data"]["note"] = note_html

            return {
                "data": {
                    "id": tp.id,
                    "text": tp.text,
                    "_level": "testPoint",
                    "_status": tp.status,
                    "_source": tp.source,
                    "_marked": tp.marked,
                    "description": tp.description
                },
                "children": children
            }

    async def _build_mindmap_snapshot(self, db: AsyncSession, session_id: str) -> Optional[Dict[str, Any]]:
        result = await db.execute(
            select(AISession).where(AISession.id == session_id)
        )
        session = result.scalar_one_or_none()
        if not session:
            return None
        try:
            mindmap = await self.get_mindmap_data(db, session.requirement_id)
            return mindmap.model_dump(by_alias=True)
        except Exception:
            return None

    async def _apply_mindmap_changes(
        self, db: AsyncSession, requirement_id: str, node_id: str,
        node_type: str, marked_node_ids: List[str], pending_data: Any
    ) -> List[Dict[str, str]]:
        """应用脑图变更，返回新建节点的ID映射 [{text, id, level}]"""
        pending_nodes = pending_data.get("adjustNodes", []) if isinstance(pending_data, dict) else []
        if not pending_nodes:
            return []
        now = datetime.utcnow()
        new_node_mappings = []
        for node in pending_nodes:
            action = node.get("action", "")
            if action == "add" and node_type == "requirement":
                tp_id = f"tp-{uuid.uuid4().hex[:8]}"
                test_point = TestPoint(
                    id=tp_id,
                    split_requirement_id=node_id,
                    text=node.get("text", ""),
                    description=node.get("description"),
                    source="AI",
                    marked=False,
                    status="pending",
                    created_at=now,
                    updated_at=now
                )
                db.add(test_point)
                new_node_mappings.append({"text": node.get("text", ""), "id": tp_id, "level": "testPoint"})
            elif action == "remove" and node_type == "requirement":
                target_id = node.get("id", "")
                if target_id and target_id not in marked_node_ids:
                    await db.execute(delete(TestPoint).where(TestPoint.id == target_id))
            elif action == "modify" and node_type == "requirement":
                target_id = node.get("id", "")
                if target_id and target_id not in marked_node_ids:
                    values = {"updated_at": now}
                    if node.get("text"):
                        values["text"] = node["text"]
                    if node.get("description") is not None:
                        values["description"] = node["description"]
                    await db.execute(update(TestPoint).where(TestPoint.id == target_id).values(**values))
            elif action == "add" and node_type == "testPoint":
                tc_id = f"tc-{uuid.uuid4().hex[:8]}"
                steps_data = node.get("steps") or []
                test_case = TestCase(
                    id=tc_id,
                    test_point_id=node_id,
                    text=node.get("text", ""),
                    case_property=node.get("case_property", "正例"),
                    pre_condition=node.get("pre_condition"),
                    steps=steps_data,
                    source="AI",
                    marked=False,
                    created_at=now,
                    updated_at=now
                )
                db.add(test_case)
                new_node_mappings.append({"text": node.get("text", ""), "id": tc_id, "level": "testCase"})
            elif action == "remove" and node_type == "testPoint":
                target_id = node.get("id", "")
                if target_id and target_id not in marked_node_ids:
                    await db.execute(delete(TestCase).where(TestCase.id == target_id))
            elif action == "modify" and node_type == "testPoint":
                target_id = node.get("id", "")
                if target_id and target_id not in marked_node_ids:
                    values = {"updated_at": now}
                    if node.get("text"):
                        values["text"] = node["text"]
                    if node.get("case_property"):
                        values["case_property"] = node["case_property"]
                    if node.get("pre_condition") is not None:
                        values["pre_condition"] = node["pre_condition"]
                    if node.get("steps") is not None:
                        values["steps"] = node["steps"]
                    await db.execute(update(TestCase).where(TestCase.id == target_id).values(**values))
        await db.commit()
        return new_node_mappings

    async def get_ai_messages(self, db: AsyncSession, session_id: str) -> List[Dict[str, Any]]:
        result = await db.execute(
            select(AIMessage).where(AIMessage.session_id == session_id).order_by(AIMessage.created_at)
        )
        messages = result.scalars().all()
        return [
            {
                "id": msg.id,
                "role": msg.role,
                "content": msg.content,
                "type": msg.msg_type or "text",
                "changeSummary": msg.change_summary,
                "pendingMindMapData": msg.pending_mindmap_data.get("nodeData") if isinstance(msg.pending_mindmap_data, dict) and "nodeData" in msg.pending_mindmap_data else msg.pending_mindmap_data,
                "adopted": msg.adopted,
                "rejected": msg.rejected,
                "createdAt": msg.created_at.isoformat() if msg.created_at else None,
            }
            for msg in messages
        ]

    async def apply_ai_adjust(self, db: AsyncSession, session_id: str, data: AIAdjustApply) -> AIAdjustApplyResponse:
        result = await db.execute(
            select(AISession).where(AISession.id == session_id)
        )
        session = result.scalar_one_or_none()
        if not session:
            raise ValueError("会话不存在")

        last_proposal_result = await db.execute(
            select(AIMessage).where(
                AIMessage.session_id == session_id,
                AIMessage.msg_type == "proposal"
            ).order_by(AIMessage.created_at.desc())
        )
        last_proposal = last_proposal_result.scalars().first()

        adjusted_data = None
        added_count = 0
        removed_count = 0
        modified_count = 0

        if last_proposal and last_proposal.pending_mindmap_data:
            pending_data = last_proposal.pending_mindmap_data
            node_data = pending_data.get("nodeData") if isinstance(pending_data, dict) else None
            adjust_nodes = pending_data.get("adjustNodes", []) if isinstance(pending_data, dict) else []

            if node_data:
                adjusted_data = node_data

            for node in adjust_nodes:
                action = node.get("action", "")
                if action == "add":
                    added_count += 1
                elif action == "remove":
                    removed_count += 1
                elif action == "modify":
                    modified_count += 1

            await self._apply_mindmap_changes(
                db, session.requirement_id, session.node_id,
                session.node_type, session.marked_node_ids or [], pending_data
            )

        await db.execute(
            update(AISession).where(AISession.id == session_id).values(status="applied")
        )
        await db.commit()

        return AIAdjustApplyResponse(
            adjustedMindMapData=adjusted_data or {},
            addedCount=added_count,
            removedCount=removed_count,
            modifiedCount=modified_count,
            preservedCount=len(data.markedTestPointTexts or [])
        )

    # ========== 异步任务 ==========
    async def start_generation(
        self, db: AsyncSession, requirement_id: str, use_knowledge_base: bool,
        task_type: str = "points_generation"
    ) -> GenerateResponse:
        """启动生成任务。task_type: points_generation / points_regeneration / cases_generation / cases_regeneration"""
        # 检查是否已有运行中的任务
        existing_task = await db.execute(
            select(Task).where(
                and_(
                    Task.requirement_id == requirement_id,
                    Task.status.in_(["pending", "running"])
                )
            ).limit(1)
        )
        if existing_task.scalar_one_or_none():
            raise ValueError("已有生成任务正在运行，请等待完成或取消后再试")

        # 根据 task_type 设置目标状态
        target_status_map = {
            "points_generation": "generating_points",
            "points_regeneration": "generating_points",
            "cases_generation": "generating_cases",
            "cases_regeneration": "generating_cases",
        }
        target_status = target_status_map.get(task_type, "generating_points")

        await db.execute(
            update(Requirement).where(Requirement.id == requirement_id).values(status=target_status)
        )

        task = Task(
            requirement_id=requirement_id,
            status="pending",
            progress=0,
            progress_text="准备生成...",
            use_knowledge_base=use_knowledge_base,
            task_type=task_type,
        )
        db.add(task)
        await db.commit()
        await db.refresh(task)

        bg_task = asyncio.create_task(
            self._run_generation(task.id, requirement_id, use_knowledge_base, task_type)
        )
        self.tasks[task.id] = bg_task

        return GenerateResponse(taskId=task.id, taskType=task_type)

    async def _run_generation(
        self, task_id: str, requirement_id: str, use_knowledge_base: bool,
        task_type: str = "points_generation"
    ):
        from app.agents.orchestrator import TestDesignOrchestrator
        from app.core.database import AsyncSessionLocal

        is_cases = task_type in ("cases_generation", "cases_regeneration")
        is_points_regen = (task_type == "points_regeneration")
        regenerate_all = (task_type == "cases_regeneration")

        orchestrator = TestDesignOrchestrator()
        async with AsyncSessionLocal() as db:
            try:
                # 检查任务是否已被取消
                result = await db.execute(
                    select(Task.status).where(Task.id == task_id)
                )
                current_status = result.scalar()
                if current_status and current_status not in ("pending", "running"):
                    logger.info(f"Task {task_id} was cancelled before execution, skipping")
                    return

                await db.execute(
                    update(Task).where(Task.id == task_id).values(
                        status="running", progress=5,
                        progress_text="正在分析需求结构..."
                    )
                )
                await db.commit()

                async def progress_callback(progress: int, text: str):
                    check = await db.execute(select(Task.status).where(Task.id == task_id))
                    if check.scalar() not in ("running", "pending"):
                        raise Exception("TASK_CANCELLED")
                    await db.execute(
                        update(Task).where(Task.id == task_id).values(
                            progress=progress, progress_text=text
                        )
                    )
                    await db.commit()

                # 根据 task_type 调用不同的方法
                if is_cases:
                    await orchestrator.run_cases(
                        db=db,
                        requirement_id=requirement_id,
                        use_knowledge_base=use_knowledge_base,
                        regenerate_all=regenerate_all,
                        progress_callback=progress_callback,
                    )
                else:
                    await orchestrator.run_points(
                        db=db,
                        requirement_id=requirement_id,
                        use_knowledge_base=use_knowledge_base,
                        regenerate_all=is_points_regen,
                        progress_callback=progress_callback,
                    )

                # 完成前再次检查，避免覆盖 cancelled/failed 状态
                final_check = await db.execute(select(Task.status).where(Task.id == task_id))
                if final_check.scalar() == "running":
                    await db.execute(
                        update(Task).where(Task.id == task_id).values(
                            status="completed",
                            progress=100,
                            progress_text="生成完成"
                        )
                    )
                    await db.commit()

            except asyncio.CancelledError:
                logger.info(f"Task {task_id} was cancelled via asyncio cancel")
                await db.execute(
                    update(Task).where(Task.id == task_id).values(
                        status="cancelled",
                        progress_text="任务已取消"
                    )
                )
                await db.commit()
            except Exception as e:
                error_msg = str(e)
                if error_msg == "TASK_CANCELLED":
                    logger.info(f"Task {task_id} was cancelled during generation")
                    return
                logger.error(f"Task {task_id} failed: {error_msg}")
                await db.execute(
                    update(Task).where(Task.id == task_id).values(
                        status="failed",
                        progress_text=f"生成失败: {error_msg}",
                        result={"error": error_msg}
                    )
                )
                # 回退 Requirement 状态为 failed
                await db.execute(
                    update(Requirement).where(Requirement.id == requirement_id).values(status="failed")
                )
                await db.commit()
            finally:
                self.tasks.pop(task_id, None)
                await orchestrator.close()

    async def get_task_status(self, db: AsyncSession, task_id: str) -> TaskStatusResponse:
        result = await db.execute(select(Task).where(Task.id == task_id))
        task = result.scalar_one_or_none()
        if not task:
            raise ValueError("任务不存在")
        return TaskStatusResponse(
            taskId=task.id,
            requirementId=task.requirement_id,
            status=task.status,
            progress=task.progress,
            progressText=task.progress_text,
            taskType=task.task_type,
        )

    async def get_active_task(self, db: AsyncSession, requirement_id: str) -> Optional[TaskStatusResponse]:
        result = await db.execute(
            select(Task).where(
                and_(
                    Task.requirement_id == requirement_id,
                    Task.status.in_(["pending", "running"])
                )
            ).order_by(Task.created_at.desc()).limit(1)
        )
        task = result.scalar_one_or_none()
        if not task:
            return None
        return TaskStatusResponse(
            taskId=task.id,
            requirementId=task.requirement_id,
            status=task.status,
            progress=task.progress,
            progressText=task.progress_text,
            taskType=task.task_type,
        )

    async def cancel_task(self, db: AsyncSession, task_id: str) -> bool:
        # 先获取任务信息，确定 task_type
        result = await db.execute(select(Task).where(Task.id == task_id))
        task = result.scalar_one_or_none()
        if not task:
            return False

        is_cases = task.task_type in ("cases_generation", "cases_regeneration") if task.task_type else False
        requirement_id = task.requirement_id

        # 更新 Task 状态
        await db.execute(
            update(Task).where(Task.id == task_id).values(
                status="cancelled",
                progress_text="任务已取消"
            )
        )

        # 数据回滚
        if is_cases:
            # 删除该需求下所有用例
            tp_subq = (
                select(TestPoint.id)
                .join(SplitRequirement)
                .where(SplitRequirement.requirement_id == requirement_id)
            ).subquery()
            await db.execute(
                delete(TestCase).where(TestCase.test_point_id.in_(select(tp_subq)))
            )
            # 恢复状态
            await db.execute(
                update(Requirement).where(Requirement.id == requirement_id).values(status="points_generated")
            )
        else:
            # 删除该需求下所有测试点（需先删用例再删测试点）
            tp_subq = (
                select(TestPoint.id)
                .join(SplitRequirement)
                .where(SplitRequirement.requirement_id == requirement_id)
            ).subquery()
            await db.execute(delete(TestCase).where(TestCase.test_point_id.in_(select(tp_subq))))
            await db.execute(
                delete(TestPoint).where(TestPoint.id.in_(select(tp_subq)))
            )
            # 恢复状态
            await db.execute(
                update(Requirement).where(Requirement.id == requirement_id).values(status="confirmed")
            )

        await db.commit()

        # 取消后台 asyncio 协程
        bg_task = self.tasks.pop(task_id, None)
        if bg_task and not bg_task.done():
            bg_task.cancel()
            try:
                await bg_task
            except asyncio.CancelledError:
                pass

        return True

    # ========== Excel导出 ==========
    async def export_excel(self, db: AsyncSession, requirement_id: str) -> bytes:
        result = await db.execute(
            select(SplitRequirement)
            .where(SplitRequirement.requirement_id == requirement_id)
            .options(selectinload(SplitRequirement.test_points).selectinload(TestPoint.test_cases))
        )
        split_reqs = result.scalars().all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"
        
        headers = ["测试用例名称", "用例类型", "前置条件", "步骤名字", "步骤描述", "步骤预期结果"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        col_widths = [30, 10, 30, 20, 30, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 2
        for sr in split_reqs:
            for tp in sr.test_points:
                for tc in tp.test_cases:
                    steps = tc.steps or []
                    if not steps:
                        steps = [{"name": "", "description": "", "stepExpectedResult": ""}]
                    
                    start_row = current_row
                    for step in steps:
                        ws.append([
                            tc.text,
                            tc.case_property,
                            tc.pre_condition or "",
                            step.get("name", ""),
                            step.get("description", ""),
                            step.get("stepExpectedResult", "")
                        ])
                        current_row += 1
                    
                    if len(steps) > 1:
                        for col in [1, 2, 3]:
                            ws.merge_cells(start_row=start_row, start_column=col, end_row=current_row - 1, end_column=col)
                            cell = ws.cell(row=start_row, column=col)
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


test_design_service = TestDesignService()
