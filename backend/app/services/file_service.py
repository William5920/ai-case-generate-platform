import os
import uuid
from datetime import datetime
from typing import Optional
from fastapi import UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.config import settings
from app.models.db_models import UploadedFile


class FileService:

    ALLOWED_EXTENSIONS = {".doc", ".docx", ".pdf", ".md", ".xlsx"}
    MAX_FILE_SIZE = 10 * 1024 * 1024

    async def upload_file(
        self,
        db: AsyncSession,
        user_id: str,
        file: UploadFile,
        purpose: str = "requirement"
    ) -> dict:
        original_filename = file.filename
        _, ext = os.path.splitext(original_filename)
        ext = ext.lower()

        if ext not in self.ALLOWED_EXTENSIONS:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 {', '.join(self.ALLOWED_EXTENSIONS)}")

        content = await file.read()
        file_size = len(content)

        if file_size > self.MAX_FILE_SIZE:
            raise ValueError(f"文件大小超过限制（最大10MB），当前文件大小: {file_size / 1024 / 1024:.1f}MB")

        file_id = f"file-{uuid.uuid4().hex[:8]}"
        upload_dir = settings.UPLOAD_DIR
        os.makedirs(upload_dir, exist_ok=True)

        save_filename = f"{file_id}{ext}"
        save_path = os.path.join(upload_dir, save_filename)

        with open(save_path, "wb") as f:
            f.write(content)

        file_type = file.content_type or "application/octet-stream"

        uploaded_file = UploadedFile(
            id=file_id,
            user_id=user_id,
            original_filename=original_filename,
            file_path=save_path,
            file_size=file_size,
            file_type=file_type,
            purpose=purpose,
            created_at=datetime.utcnow()
        )
        db.add(uploaded_file)
        await db.commit()
        await db.refresh(uploaded_file)

        return {
            "fileId": file_id,
            "fileName": original_filename,
            "fileSize": file_size,
            "fileType": file_type,
            "uploadedAt": uploaded_file.created_at.isoformat() + "Z"
        }

    async def get_file_info(self, db: AsyncSession, file_id: str) -> Optional[dict]:
        result = await db.execute(select(UploadedFile).where(UploadedFile.id == file_id))
        uploaded_file = result.scalar_one_or_none()
        if not uploaded_file:
            return None
        return {
            "fileId": uploaded_file.id,
            "fileName": uploaded_file.original_filename,
            "fileSize": uploaded_file.file_size,
            "fileType": uploaded_file.file_type
        }

    async def parse_file_content(self, db: AsyncSession, file_id: str) -> Optional[str]:
        result = await db.execute(select(UploadedFile).where(UploadedFile.id == file_id))
        uploaded_file = result.scalar_one_or_none()
        if not uploaded_file:
            return None

        file_path = uploaded_file.file_path
        _, ext = os.path.splitext(file_path)
        ext = ext.lower()

        if ext == ".md":
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()
        elif ext == ".docx":
            return await self._parse_docx(file_path)
        elif ext == ".xlsx":
            return await self._parse_xlsx(file_path)
        elif ext == ".pdf":
            return await self._parse_pdf(file_path)
        elif ext == ".doc":
            return await self._parse_doc(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

    async def _parse_docx(self, file_path: str) -> str:
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n".join(paragraphs)
        except Exception as e:
            raise ValueError(f"docx文件解析失败: {str(e)}")

    async def _parse_xlsx(self, file_path: str) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            content_parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        content_parts.append(row_text)
            wb.close()
            return "\n".join(content_parts)
        except Exception as e:
            raise ValueError(f"xlsx文件解析失败: {str(e)}")

    async def _parse_pdf(self, file_path: str) -> str:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n".join(text_parts)
        except Exception as e:
            raise ValueError(f"pdf文件解析失败: {str(e)}")

    async def _parse_doc(self, file_path: str) -> str:
        raise ValueError("暂不支持.doc格式，请转换为.docx后上传")


file_service = FileService()
